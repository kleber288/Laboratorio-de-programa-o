"""Carrinho, finalização, histórico e exportação de vendas."""

from io import BytesIO

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font

from models import produto as produto_model
from models import venda as venda_model

venda_bp = Blueprint("venda", __name__, url_prefix="/vendas")


def _carrinho():
    return session.setdefault("carrinho", [])


@venda_bp.get("")
def listar():
    return render_template("vendas.html", vendas=venda_model.listar())


@venda_bp.get("/nova")
def nova():
    carrinho = _carrinho()
    total = sum(item["preco_centavos"] * item["quantidade"] for item in carrinho)
    return render_template("nova_venda.html", produtos=produto_model.listar(), carrinho=carrinho, total=total)


@venda_bp.post("/carrinho/adicionar")
def adicionar():
    try:
        produto_id = int(request.form.get("produto_id", ""))
        quantidade = int(request.form.get("quantidade", ""))
    except ValueError:
        produto_id, quantidade = 0, 0
    produto = produto_model.buscar(produto_id)
    if produto is None or quantidade <= 0:
        flash("Selecione um produto e uma quantidade válida.", "erro")
    elif quantidade > produto["estoque"]:
        flash("A quantidade supera o estoque disponível.", "erro")
    else:
        carrinho = _carrinho()
        existente = next((item for item in carrinho if item["produto_id"] == produto_id), None)
        nova_quantidade = quantidade + (existente["quantidade"] if existente else 0)
        if nova_quantidade > produto["estoque"]:
            flash("A quantidade total no carrinho supera o estoque.", "erro")
            return redirect(url_for("venda.nova"))
        if existente:
            existente["quantidade"] = nova_quantidade
        else:
            carrinho.append({"produto_id": produto["id"], "nome": produto["nome"], "preco_centavos": produto["preco_centavos"], "quantidade": quantidade})
        session["carrinho"] = carrinho
        flash("Produto adicionado à venda.", "sucesso")
    return redirect(url_for("venda.nova"))


@venda_bp.post("/carrinho/<int:indice>/remover")
def remover(indice):
    carrinho = _carrinho()
    if indice < 0 or indice >= len(carrinho):
        abort(404)
    carrinho.pop(indice)
    session["carrinho"] = carrinho
    return redirect(url_for("venda.nova"))


@venda_bp.post("/carrinho/limpar")
def limpar():
    session.pop("carrinho", None)
    flash("Carrinho esvaziado.", "sucesso")
    return redirect(url_for("venda.nova"))


@venda_bp.post("/finalizar")
def finalizar():
    carrinho = _carrinho()
    cliente = request.form.get("cliente", "").strip()
    if not carrinho:
        flash("O carrinho está vazio.", "erro")
        return redirect(url_for("venda.nova"))
    if len(cliente) < 2:
        flash("Informe o nome do cliente.", "erro")
        return redirect(url_for("venda.nova"))
    try:
        venda_id = venda_model.finalizar(session["usuario_id"], cliente, carrinho)
    except ValueError as erro:
        flash(str(erro), "erro")
        return redirect(url_for("venda.nova"))
    session.pop("carrinho", None)
    flash("Venda finalizada com sucesso.", "sucesso")
    return redirect(url_for("venda.detalhes", venda_id=venda_id))


@venda_bp.get("/<int:venda_id>")
def detalhes(venda_id):
    venda = venda_model.buscar(venda_id)
    if venda is None:
        abort(404)
    return render_template("detalhes_venda.html", venda=venda, itens=venda_model.listar_itens(venda_id))


@venda_bp.get("/exportar")
def exportar():
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"
    ws.append(["Código", "Data e hora", "Cliente", "Responsável", "Total (R$)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for venda in venda_model.listar():
        ws.append([venda["id"], venda["criada_em"], venda["cliente"], venda["usuario_nome"], venda["total_centavos"] / 100])
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 24
    for cell in ws["E"][1:]:
        cell.number_format = 'R$ #,##0.00'
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return send_file(arquivo, as_attachment=True, download_name="vendas_papelaria.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
