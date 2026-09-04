"""Controller do CRUD de produtos, pesquisa e exportação para Excel."""

# Um arquivo em memória evita criar planilhas temporárias no disco do servidor.
from io import BytesIO

# ``abort`` entrega erros HTTP ao Flask; os demais imports recebem dados,
# renderizam Views, criam mensagens ou respostas de redirecionamento/download.
from flask import (
    Blueprint,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)

from openpyxl import Workbook
from openpyxl.styles import Font

from models import produto as produto_model
# O projeto importa ``utils`` embora o arquivo distribuído se chame ``util.py``;
# a divergência existente foi preservada e pode causar erro de importação.
from utils import moeda_para_centavos


produto_bp = Blueprint(
    "produto",
    __name__,
    url_prefix="/produtos"
)


def _dados_formulario():
    """Lê, converte e valida os campos comuns aos formulários de produto.

    Parâmetros:
        Nenhum explícito; usa ``request.form`` com os campos ``nome``,
        ``fabricante``, ``unidade``, ``estoque`` e ``preco`` definidos pelos
        atributos HTML ``name`` de ``produto_form.html``.

    Retorno:
        Tupla ``(nome, fabricante, unidade, preco_centavos, estoque)`` na ordem
        esperada pelos métodos ``criar`` e ``atualizar`` do Model.

    Efeitos colaterais:
        Nenhum; apenas lê a requisição atual.

    Possíveis erros:
        ``ValueError`` em inteiros, moeda ou validação; ``TypeError`` pode surgir
        de dados incompatíveis e é tratado pelas rotas chamadoras.
    """
    # ``strip`` evita aceitar campos compostos somente de espaços.
    nome = request.form.get("nome", "").strip()
    fabricante = request.form.get("fabricante", "").strip()
    unidade = request.form.get("unidade", "").strip()
    # Converte na borda HTTP: Models recebem tipos prontos para persistência.
    estoque = int(request.form.get("estoque", "0"))
    preco = moeda_para_centavos(request.form.get("preco", ""))

    # A validação de servidor complementa ``required`` e ``min`` do HTML.
    if (
        not nome
        or not fabricante
        or not unidade
        or estoque < 0
        or preco < 0
    ):
        raise ValueError(
            "Preencha todos os campos com valores válidos."
        )

    return nome, fabricante, unidade, preco, estoque


@produto_bp.get("")
def listar():
    """Lista/pesquisa produtos em ``GET /produtos``.

    Parâmetros:
        Nenhum explícito; o termo opcional ``busca`` vem da query string por
        ``request.args`` (por exemplo, ``/produtos?busca=cafe``).

    Retorno:
        HTML de ``produtos.html`` com as linhas do Model e o termo pesquisado.

    Efeitos colaterais:
        Apenas consulta o banco.

    Possíveis erros:
        Falhas SQLite/Jinja são propagadas. O middleware exige login antes.
    """
    busca = request.args.get("busca", "")

    # O valor de ``busca`` volta à View para permanecer visível no campo.
    return render_template(
        "produtos.html",
        produtos=produto_model.listar(busca),
        busca=busca,
    )


@produto_bp.route("/novo", methods=("GET", "POST"))
def novo():
    """Exibe em GET e cria em POST um produto em ``/produtos/novo``.

    Parâmetros:
        Nenhum explícito; o POST é lido por ``_dados_formulario``.

    Retorno:
        Redirecionamento à listagem após sucesso ou o formulário com
        ``produto=None`` em GET/falha.

    Efeitos colaterais:
        O Model executa INSERT/commit e ``flash`` agenda uma mensagem.

    Possíveis erros:
        ``ValueError`` e ``TypeError`` de conversão são exibidos como validação;
        outras falhas do banco são propagadas.
    """
    if request.method == "POST":
        try:
            # O asterisco desempacota a tupla na assinatura de ``criar``.
            produto_model.criar(*_dados_formulario())

            flash("Produto cadastrado.", "sucesso")

            # Post/Redirect/Get evita repetir o INSERT ao atualizar o navegador.
            return redirect(url_for("produto.listar"))

        except (ValueError, TypeError):
            flash(
                "Preencha preço e estoque com valores válidos.",
                "erro",
            )

    return render_template(
        "produto_form.html",
        produto=None,
    )


@produto_bp.route(
    "/<int:produto_id>/editar",
    methods=("GET", "POST"),
)
def editar(produto_id):
    """Exibe ou atualiza um produto em ``/produtos/<id>/editar``.

    Parâmetros:
        produto_id: Inteiro extraído da própria URL pelo conversor ``<int:...>``.

    Retorno:
        404 se não existir; redirecionamento após POST válido; caso contrário,
        HTML de ``produto_form.html`` com a linha atual em ``produto``.

    Efeitos colaterais:
        Consulta sempre e, no POST válido, executa UPDATE/commit e cria flash.

    Possíveis erros:
        Conversões inválidas são tratadas; falhas SQLite/Jinja são propagadas.
    """
    # Buscar antes serve tanto para preencher o GET quanto para validar o id.
    produto = produto_model.buscar(produto_id)

    if produto is None:
        # ``abort`` aciona o tratador 404 registrado em ``app.py``.
        abort(404)

    if request.method == "POST":
        try:
            produto_model.atualizar(
                produto_id,
                *_dados_formulario(),
            )

            flash("Produto atualizado.", "sucesso")

            return redirect(url_for("produto.listar"))

        except (ValueError, TypeError):
            flash(
                "Preencha preço e estoque com valores válidos.",
                "erro",
            )

    return render_template(
        "produto_form.html",
        produto=produto,
    )


@produto_bp.post("/<int:produto_id>/excluir")
def excluir(produto_id):
    """Exclui por POST o produto em ``/produtos/<id>/excluir``.

    Parâmetros:
        produto_id: Inteiro capturado da URL.

    Retorno:
        Redirecionamento à listagem ou resposta 404 para id ausente.

    Efeitos colaterais:
        Executa SELECT, DELETE, commit e mensagem flash. Pela relação
        ``ON DELETE SET NULL``, itens de compras antigas deveriam manter a cópia
        de nome e preço, perdendo apenas o vínculo ao produto.

    Possíveis erros:
        Restrições ou falhas SQLite são propagadas. O decorator POST evita
        exclusões acionadas por navegação GET.
    """
    if produto_model.buscar(produto_id) is None:
        abort(404)

    produto_model.excluir(produto_id)

    flash(
        "Produto excluído. Compras antigas preservam seu nome e preço.",
        "sucesso",
    )

    return redirect(url_for("produto.listar"))


@produto_bp.get("/exportar")
def exportar():
    """Gera e baixa a planilha em ``GET /produtos/exportar``.

    Parâmetros:
        Nenhum.

    Retorno:
        Resposta ``send_file`` com ``produtos.xlsx``, MIME do formato XLSX e
        cabeçalho de anexo.

    Efeitos colaterais:
        Consulta produtos e cria um ``Workbook`` somente em memória.

    Possíveis erros:
        Falhas do banco, openpyxl, memória ou serialização são propagadas.
    """
    # ``Workbook`` cria uma pasta com uma worksheet inicial ativa.
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # ``append`` preenche a próxima linha; esta primeira linha é o cabeçalho.
    ws.append([
        "Código",
        "Nome",
        "Fabricante",
        "Unidade",
        "Preço (R$)",
        "Estoque",
    ])

    # ``ws[1]`` seleciona todas as células da primeira linha para aplicar negrito.
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Cada ``sqlite3.Row`` vira uma linha; centavos / 100 gera valor numérico
    # para o Excel, permitindo cálculos e formatação monetária real.
    for produto in produto_model.listar():
        ws.append([
            produto["id"],
            produto["nome"],
            produto["fabricante"],
            produto["unidade"],
            produto["preco_centavos"] / 100,
            produto["estoque"],
        ])

    # Larguras melhoram a leitura sem modificar o conteúdo das células.
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22

    # A fatia ignora o cabeçalho e aplica o formato de moeda às células de preço.
    for cell in ws["E"][1:]:
        cell.number_format = 'R$ #,##0.00'

    # ``BytesIO`` funciona como arquivo binário em RAM.
    arquivo = BytesIO()

    wb.save(arquivo)
    # Após ``save``, o cursor está no fim; ``seek(0)`` volta ao início para que
    # ``send_file`` não produza um download vazio.
    arquivo.seek(0)

    # ``as_attachment`` orienta o navegador a baixar em vez de exibir.
    return send_file(
        arquivo,
        as_attachment=True,
        download_name="produtos.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
