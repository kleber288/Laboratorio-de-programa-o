"""Controller do carrinho, finalização, histórico e exportação de compras."""

# Buffer binário que recebe a planilha sem gravar arquivo permanente.
from io import BytesIO

# Os proxies request/session pertencem à requisição; funções de resposta ligam
# este Controller às Views, mensagens, erros HTTP, redirecionamentos e downloads.
from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font

from models import compra as compra_model
from models import produto as produto_model


# O prefixo compõe todas as URLs deste domínio, como ``/compras/nova``.
compra_bp = Blueprint("compra", __name__, url_prefix="/compras")


def _carrinho():
    """Obtém o carrinho persistido na sessão do visitante autenticado.

    Parâmetros:
        Nenhum.

    Retorno:
        Lista de dicionários com produto, nome, preço e quantidade.

    Efeitos colaterais:
        ``setdefault`` cria a chave ``carrinho`` com lista vazia se ainda não
        existir; a sessão será serializada no cookie assinado pelo Flask.

    Possíveis erros:
        Fora de contexto de requisição, ``session`` gera erro. Todas as rotas de
        compra usam este formato compartilhado.
    """
    return session.setdefault("carrinho", [])


@compra_bp.get("")
def listar():
    """Atende ``GET /compras`` com o histórico completo.

    Não recebe parâmetros. Retorna ``compras.html`` com ``compras`` vindas do
    Model; o efeito é apenas SELECT. Erros SQLite/Jinja são propagados, e o
    middleware de ``app.py`` bloqueia visitantes não autenticados.
    """
    return render_template("compras.html", compras=compra_model.listar())


@compra_bp.get("/nova")
def nova():
    """Atende ``GET /compras/nova`` e apresenta catálogo, carrinho e total.

    Não recebe parâmetros. Retorna ``nova_compra.html`` com ``produtos``,
    ``carrinho`` e ``total``; pode criar a chave de sessão do carrinho e consulta
    produtos. Erros de sessão, banco ou template são propagados.
    """
    carrinho = _carrinho()
    # O total deriva do preço congelado no carrinho vezes a quantidade de cada item.
    total = sum(i["preco_centavos"] * i["quantidade"] for i in carrinho)
    return render_template("nova_compra.html", produtos=produto_model.listar(), carrinho=carrinho, total=total)


@compra_bp.post("/carrinho/adicionar")
def adicionar():
    """Adiciona quantidade em ``POST /compras/carrinho/adicionar``.

    Parâmetros:
        Nenhum explícito; ``produto_id`` e ``quantidade`` vêm de ``request.form``
        e dos ``name`` correspondentes em ``nova_compra.html``.

    Retorno:
        Sempre redireciona para ``compra.nova``.

    Efeitos colaterais:
        Consulta o produto, cria/atualiza a sessão e registra mensagem flash.

    Possíveis erros:
        ``ValueError`` de conversão é normalizado para valores inválidos; falhas
        do banco/sessão são propagadas. O estoque é validado novamente na
        finalização para cobrir mudanças concorrentes.
    """
    try:
        # Campos de formulário são texto; a conversão explícita valida números inteiros.
        produto_id = int(request.form.get("produto_id", ""))
        quantidade = int(request.form.get("quantidade", ""))
    except ValueError:
        produto_id, quantidade = 0, 0
    produto = produto_model.buscar(produto_id)
    # Primeiro valida existência e positividade, depois o limite do estoque atual.
    if produto is None or quantidade <= 0:
        flash("Selecione um produto e uma quantidade válida.", "erro")
    elif quantidade > produto["estoque"]:
        flash("A quantidade supera o estoque disponível.", "erro")
    else:
        carrinho = _carrinho()
        # ``next(..., None)`` procura o mesmo produto sem criar entrada duplicada.
        existente = next((i for i in carrinho if i["produto_id"] == produto_id), None)
        nova_quantidade = quantidade + (existente["quantidade"] if existente else 0)
        # A soma precisa ser validada, pois cada adição isolada pode caber no saldo.
        if nova_quantidade > produto["estoque"]:
            flash("A quantidade total no carrinho supera o estoque.", "erro")
            return redirect(url_for("compra.nova"))
        if existente:
            existente["quantidade"] = nova_quantidade
        else:
            carrinho.append({"produto_id": produto["id"], "nome": produto["nome"],
                             "preco_centavos": produto["preco_centavos"], "quantidade": quantidade})
        # Reatribuir informa ao mecanismo de sessão que o valor mutável mudou.
        session["carrinho"] = carrinho
        flash("Produto adicionado ao carrinho.", "sucesso")
    return redirect(url_for("compra.nova"))


@compra_bp.post("/carrinho/<int:indice>/remover")
def remover(indice):
    """Remove por posição em ``POST /compras/carrinho/<indice>/remover``.

    ``indice`` é um inteiro vindo da URL e corresponde a ``loop.index0`` da View.
    Retorna redirecionamento para a compra nova; altera a lista da sessão. Índice
    fora dos limites chama ``abort(404)`` e usa o tratador de ``app.py``.
    """
    carrinho = _carrinho()
    if indice < 0 or indice >= len(carrinho):
        abort(404)
    carrinho.pop(indice)
    session["carrinho"] = carrinho
    return redirect(url_for("compra.nova"))


@compra_bp.post("/carrinho/limpar")
def limpar():
    """Esvazia o carrinho em ``POST /compras/carrinho/limpar``.

    Não recebe parâmetros. Remove a chave da sessão, cria flash e redireciona
    para ``compra.nova``. ``pop(..., None)`` não falha se a chave não existir;
    problemas de sessão podem ser propagados.
    """
    session.pop("carrinho", None)
    flash("Carrinho esvaziado.", "sucesso")
    return redirect(url_for("compra.nova"))


@compra_bp.post("/finalizar")
def finalizar():
    """Conclui a venda em ``POST /compras/finalizar``.

    Não recebe parâmetros explícitos; usa ``usuario_id`` e ``carrinho`` da
    sessão. Retorna redirecionamento à nova compra em erro validável ou aos
    detalhes em sucesso. O Model cria compra/itens, verifica e baixa estoque em
    transação; depois o Controller limpa o carrinho e cria flash.

    ``ValueError`` (produto removido/estoque insuficiente) é mostrado ao usuário;
    outros erros são propagados após o rollback feito pelo Model.
    """
    carrinho = _carrinho()
    if not carrinho:
        flash("O carrinho está vazio.", "erro")
        return redirect(url_for("compra.nova"))
    try:
        # A identidade da sessão liga a compra ao responsável autenticado.
        compra_id = compra_model.finalizar(session["usuario_id"], carrinho)
    except ValueError as erro:
        flash(str(erro), "erro")
        return redirect(url_for("compra.nova"))
    # Só limpa após o commit bem-sucedido; em falha, o usuário pode revisar itens.
    session.pop("carrinho", None)
    flash("Compra finalizada com sucesso.", "sucesso")
    return redirect(url_for("compra.detalhes", compra_id=compra_id))


@compra_bp.get("/<int:compra_id>")
def detalhes(compra_id):
    """Exibe ``GET /compras/<compra_id>`` como comprovante imprimível.

    ``compra_id`` é convertido da URL. Retorna ``detalhes_compra.html`` com o
    cabeçalho ``compra`` e seus ``itens``; realiza dois SELECTs. Ausência gera
    404, e outras falhas de banco/template são propagadas.
    """
    compra = compra_model.buscar(compra_id)
    if compra is None:
        abort(404)
    return render_template("detalhes_compra.html", compra=compra,
                           itens=compra_model.listar_itens(compra_id))


@compra_bp.get("/exportar")
def exportar():
    """Gera o download de ``compras.xlsx`` em ``GET /compras/exportar``.

    Não recebe parâmetros. Retorna resposta XLSX como anexo via ``send_file``;
    consulta o histórico e monta um ``Workbook`` em memória. Pode propagar erros
    SQLite, de memória ou do openpyxl.
    """
    # O Workbook nasce com uma worksheet ativa, aqui renomeada para o domínio.
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"
    # ``append`` escreve cabeçalho e, no laço, uma linha por compra.
    ws.append(["Código", "Data e hora", "Responsável", "Total (R$)"])
    # Células da primeira linha recebem fonte em negrito.
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Dividir centavos por 100 entrega número que o Excel pode somar/formatar.
    for c in compra_model.listar():
        ws.append([c["id"], c["criada_em"], c["usuario_nome"], c["total_centavos"] / 100])
    # Ajustes de largura favorecem datas e nomes sem alterar os dados.
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 25
    # A fatia pula o cabeçalho e aplica máscara monetária aos totais.
    for cell in ws["D"][1:]:
        cell.number_format = 'R$ #,##0.00'
    # ``BytesIO`` evita arquivo temporário; ``save`` deixa o cursor no final.
    arquivo = BytesIO()
    wb.save(arquivo)
    # Voltar à posição zero é essencial para ``send_file`` ler todo o XLSX.
    arquivo.seek(0)
    # MIME identifica o formato e ``as_attachment`` solicita download ao navegador.
    return send_file(arquivo, as_attachment=True, download_name="compras.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
