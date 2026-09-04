"""Testes de integração: exercitam rotas, sessão, SQLite, Views e XLSX juntos."""

# ``io.BytesIO`` permite entregar ao openpyxl os bytes baixados sem arquivo físico.
import io
import sqlite3

from openpyxl import load_workbook

from app import create_app


def login(client):
    """Autentica o administrador padrão para reutilização nos testes.

    Recebe o cliente Flask e retorna a resposta final após seguir redirects.
    Como efeito, cria a sessão autenticada do cliente; pode expor falhas da rota,
    schema ou usuário semeado.
    """
    # ``follow_redirects`` faz a asserção enxergar a página de destino, não só 302.
    return client.post("/login", data={"usuario": "admin", "senha": "admin123"}, follow_redirects=True)


def test_middleware_redireciona_visitante(client):
    """Confirma que uma rota privada redireciona visitante ao login.

    Usa ``client`` e não retorna valor. Realiza GET sem sessão; os asserts falham
    com mensagem do pytest se o status não for 302 ou o Location não apontar ao
    login. Assim verifica o ``before_request`` de ``app.py``.
    """
    resposta = client.get("/produtos")
    # 302 indica redirecionamento; Location mostra o destino HTTP.
    assert resposta.status_code == 302
    assert "/login" in resposta.headers["Location"]


def test_fluxo_completo(client, app):
    """Verifica login, cadastro de produto, carrinho, compra e estoque.

    Recebe cliente e aplicação configurada. Não retorna valor; cria dados no
    SQLite temporário. Pode falhar em qualquer integração Controller/Model/View.
    Os asserts conferem mensagens/totais visíveis e valores persistidos.
    """
    resposta = login(client)
    assert "Olá, Administrador" in resposta.get_data(as_text=True)

    # Os nomes do dicionário reproduzem os atributos ``name`` do formulário HTML.
    resposta = client.post("/produtos/novo", data={
        "nome": "Café", "fabricante": "Serra", "unidade": "pacote",
        "preco": "18,90", "estoque": "10",
    }, follow_redirects=True)
    assert "Produto cadastrado" in resposta.get_data(as_text=True)

    resposta = client.post("/compras/carrinho/adicionar", data={
        "produto_id": "1", "quantidade": "2"
    }, follow_redirects=True)
    assert "R$ 37,80" in resposta.get_data(as_text=True)

    resposta = client.post("/compras/finalizar", follow_redirects=True)
    pagina = resposta.get_data(as_text=True)
    assert "Compra #1" in pagina
    assert "R$ 37,80" in pagina

    # Abre uma conexão independente para verificar o estado real persistido; o
    # context manager confirma/fecha a conexão ao sair do bloco.
    with sqlite3.connect(app.config["DATABASE"]) as db:
        assert db.execute("SELECT estoque FROM produtos WHERE id=1").fetchone()[0] == 8
        assert db.execute("SELECT total_centavos FROM compras WHERE id=1").fetchone()[0] == 3780


def test_exportacoes_excel(client):
    """Valida status, MIME e nome da worksheet nas duas exportações.

    Recebe o cliente e cria sessão por ``login``. Não grava arquivos: abre os
    bytes das respostas em memória. Os asserts cobrem produtos e compras; erros
    das rotas ou um XLSX inválido fazem o teste falhar.
    """
    login(client)
    # A tabela de casos evita duplicar a mesma sequência de verificações.
    for rota, titulo in [("/produtos/exportar", "Produtos"), ("/compras/exportar", "Compras")]:
        resposta = client.get(rota)
        assert resposta.status_code == 200
        assert resposta.headers["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # ``load_workbook`` também comprova que a resposta é um pacote XLSX legível.
        workbook = load_workbook(io.BytesIO(resposta.data))
        assert workbook.active.title == titulo


def test_cadastro_e_login_novo_usuario(client):
    """Comprova que um cadastro novo pode autenticar-se em seguida.

    Recebe o cliente, persiste um usuário no banco temporário e mantém cookies
    entre requisições. Não retorna valor; asserts procuram as mensagens finais e
    revelam falhas de hash, Model, sessão ou templates.
    """
    resposta = client.post("/cadastro", data={
        "nome": "Maria Silva", "usuario": "maria", "senha": "senha123"
    }, follow_redirects=True)
    assert "Cadastro realizado" in resposta.get_data(as_text=True)
    resposta = client.post("/login", data={"usuario": "maria", "senha": "senha123"}, follow_redirects=True)
    assert "Olá, Maria Silva" in resposta.get_data(as_text=True)


def test_exclusao_e_edicao_exigem_post_ou_formulario(client):
    """Garante que ações mutáveis não aceitam requisições GET.

    Após autenticar o cliente, envia GET para exclusão e logout. Não retorna
    valor nem deve mudar dados/sessão. O status 405 prova que os decorators
    restringem os métodos HTTP e protege contra mutação por simples navegação.
    """
    login(client)
    assert client.get("/produtos/1/excluir").status_code == 405
    assert client.get("/logout").status_code == 405
